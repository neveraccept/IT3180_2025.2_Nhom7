# -*- coding: utf-8 -*-
"""
Sinh file Excel mẫu hoá đơn điện/nước/internet để Admin import qua
POST /api/utility-bills/import (chức năng "Nhập hoá đơn hàng loạt").

InitialDataConfig chọn NGẪU NHIÊN 100 hộ trong số các căn ở. Vì không biết trước hộ nào
được chọn, file này liệt kê TẤT CẢ mã hộ có thể có (145 hộ = tầng 6..29 mỗi tầng 6 căn +
penthouse PH30-01) => HK-A06-01 ... HK-A29-06, HK-PH30-01.
Khi import, dòng nào trỏ tới hộ chưa tồn tại sẽ được hệ thống TỰ ĐỘNG BỎ QUA.

Cột phải khớp IMPORT_HEADERS trong UtilityBillService:
  Mã hộ | Loại (DIEN/NUOC/INTERNET) | Tháng | Năm | Chỉ số cũ | Chỉ số mới | Số tiền (Internet)
- Điện/Nước: điền chỉ số cũ & mới; tiền tự tính = (mới - cũ) * đơn giá SystemConfig.
- Internet: điền cột "Số tiền (Internet)"; chỉ số để trống.
"""
import os
import random

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(20260608)  # cố định để file tái lập được

HEADERS = ["Mã hộ", "Loại (DIEN/NUOC/INTERNET)", "Tháng", "Năm",
           "Chỉ số cũ", "Chỉ số mới", "Số tiền (Internet)"]

MONTH = 6
YEAR = 2026
INTERNET_PRICES = [200000, 250000, 300000]


def household_codes():
    """Tất cả mã hộ có thể có giống InitialDataConfig: tầng 6..29 (6 căn/tầng) + PH30-01."""
    codes = []
    for floor in range(6, 30):              # tầng 6..29
        for i in range(1, 7):               # 6 căn mỗi tầng
            codes.append("HK-A%02d-%02d" % (floor, i))
    codes.append("HK-PH30-01")              # penthouse
    return codes


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoá đơn điện nước"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    center = Alignment(horizontal="center")

    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    row = 2
    # Chỉ số luỹ kế giả lập theo từng hộ để dữ liệu trông thực tế hơn.
    elec_meter = {}
    water_meter = {}
    for code in household_codes():
        elec_meter.setdefault(code, random.randint(1500, 6000))
        water_meter.setdefault(code, random.randint(120, 900))

        # --- ĐIỆN ---
        e_old = elec_meter[code]
        e_new = e_old + random.randint(80, 420)     # ~80-420 kWh/tháng
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value="DIEN")
        ws.cell(row=row, column=3, value=MONTH)
        ws.cell(row=row, column=4, value=YEAR)
        ws.cell(row=row, column=5, value=e_old)
        ws.cell(row=row, column=6, value=e_new)
        row += 1

        # --- NƯỚC ---
        w_old = water_meter[code]
        w_new = w_old + random.randint(5, 45)       # ~5-45 m3/tháng
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value="NUOC")
        ws.cell(row=row, column=3, value=MONTH)
        ws.cell(row=row, column=4, value=YEAR)
        ws.cell(row=row, column=5, value=w_old)
        ws.cell(row=row, column=6, value=w_new)
        row += 1

        # --- INTERNET ---
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value="INTERNET")
        ws.cell(row=row, column=3, value=MONTH)
        ws.cell(row=row, column=4, value=YEAR)
        ws.cell(row=row, column=7, value=random.choice(INTERNET_PRICES))
        row += 1

    widths = [14, 26, 8, 8, 12, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hoa-don-dien-nuoc.xlsx")
    wb.save(out)
    print("Created:", out)
    print("Data rows:", row - 2, "(", (row - 2) // 3, "households x 3 types )")


if __name__ == "__main__":
    main()
